"""
Формирование выгрузок в Excel для админки.
"""
from __future__ import annotations

from io import BytesIO
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Document, Transaction, User


async def build_users_xlsx(session: AsyncSession) -> bytes:
    """Выгрузка пользователей: tg_id, username, имя, дата регистрации, беспл./платн. лимиты, кол-во документов."""
    from sqlalchemy.orm import selectinload
    result = await session.execute(select(User).options(selectinload(User.balance)))
    users = result.scalars().all()
    docs_count_q = select(Document.user_id, func.count(Document.id).label("cnt")).group_by(Document.user_id)
    docs_result = await session.execute(docs_count_q)
    docs_map = {row[0]: row[1] for row in docs_result.fetchall()}
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["tg_id", "username", "first_name", "last_name", "created_at", "free_remaining", "purchased", "documents_count", "is_banned"]
    ws.append(headers)
    for u in users:
        ws.append([
            u.tg_id,
            u.username or "",
            u.first_name or "",
            u.last_name or "",
            u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
            u.free_limits_remaining,
            u.balance.purchased_credits if u.balance else 0,
            docs_map.get(u.id, 0),
            "да" if getattr(u, "is_banned", False) else "нет",
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def build_transactions_xlsx(session: AsyncSession) -> bytes:
    """Выгрузка транзакций: дата, user_id, tg_id, сумма, валюта, статус, описание."""
    result = await session.execute(
        select(Transaction, User.tg_id)
        .join(User, Transaction.user_id == User.id)
        .order_by(Transaction.created_at.desc())
    )
    rows = result.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    ws.append(["id", "created_at", "user_id", "tg_id", "amount", "currency", "status", "yookassa_payment_id", "description"])
    for txn, tg_id in rows:
        ws.append([
            txn.id,
            txn.created_at.strftime("%Y-%m-%d %H:%M") if txn.created_at else "",
            txn.user_id,
            tg_id,
            float(txn.amount),
            txn.currency,
            txn.status,
            txn.yookassa_payment_id or "",
            txn.description or "",
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def build_summary_xlsx(session: AsyncSession) -> bytes:
    """Сводка: итоги по пользователям, документам, выручке."""
    total_users = await session.scalar(select(func.count(User.id)))
    total_docs = await session.scalar(select(func.count(Document.id)))
    total_paid = await session.scalar(
        select(func.coalesce(func.sum(Transaction.amount), 0)).where(Transaction.status == "succeeded")
    ) or 0
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Показатель", "Значение"])
    ws.append(["Всего пользователей", total_users])
    ws.append(["Обработано документов", total_docs])
    ws.append(["Выручка (успешные платежи), ₽", float(total_paid)])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
